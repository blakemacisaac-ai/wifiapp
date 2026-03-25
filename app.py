import os
import requests
import psycopg2
import psycopg2.extras
import threading
from datetime import datetime, timezone
from functools import wraps
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, jsonify, g, send_file
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")

# ── Config ───────────────────────────────────────────────────
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
MERAKI_API_KEY = os.environ.get("MERAKI_API_KEY", "")
MERAKI_NET_ID  = os.environ.get("MERAKI_NET_ID", "")
DATABASE_URL   = os.environ.get("DATABASE_URL", "")
MERAKI_BASE    = "https://api.meraki.com/api/v1"
SLACK_WEBHOOK  = os.environ.get("SLACK_WEBHOOK", "")

# Tier → true kbps using 1024-based conversion so Meraki reports exact Mbps
# 25 Mbps = 25 * 1024 = 25600 kbps, etc.
TIERS = {
    "1": {"label": "Tier 1 — Standard",  "mbps": 25,  "kbps": 25600},
    "2": {"label": "Tier 2 — Enhanced",  "mbps": 50,  "kbps": 51200},
    "3": {"label": "Tier 3 — Premium",   "mbps": 100, "kbps": 102400},
}

# ── Database ─────────────────────────────────────────────────
def get_db():
    db = getattr(g, "_database", None)
    if db is None:
        db = g._database = psycopg2.connect(
            DATABASE_URL,
            cursor_factory=psycopg2.extras.RealDictCursor
        )
    return db

@app.teardown_appcontext
def close_db(exception):
    db = getattr(g, "_database", None)
    if db is not None:
        db.close()

def get_scheduler_db():
    """Separate DB connection for the scheduler thread (not Flask g)."""
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)

def init_db():
    with app.app_context():
        db = get_db()
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS wifi_requests (
                id               SERIAL PRIMARY KEY,
                conf_name        TEXT NOT NULL,
                ssid             TEXT NOT NULL,
                password         TEXT NOT NULL,
                start_date       TEXT NOT NULL,
                end_date         TEXT NOT NULL,
                notes            TEXT,
                tier             TEXT DEFAULT '1',
                status           TEXT DEFAULT 'pending',
                slot             INTEGER,
                error_msg        TEXT,
                submitted_at     TIMESTAMP DEFAULT NOW(),
                pushed_at        TIMESTAMP,
                enable_at        TIMESTAMP,
                disable_at       TIMESTAMP,
                schedule_status  TEXT DEFAULT 'unscheduled'
            )
        """)
        # Safe migrations for existing DBs
        for col, defn in [
            ("tier",            "TEXT DEFAULT '1'"),
            ("enable_at",       "TIMESTAMP"),
            ("disable_at",      "TIMESTAMP"),
            ("schedule_status", "TEXT DEFAULT 'unscheduled'"),
            ("archived_at",     "TIMESTAMP"),
        ]:
            cur.execute(f"ALTER TABLE wifi_requests ADD COLUMN IF NOT EXISTS {col} {defn}")
        db.commit()
        cur.close()

# ── Auth ──────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

# ── Slack notifications ───────────────────────────────────
def send_slack(msg):
    if not SLACK_WEBHOOK:
        return
    try:
        requests.post(SLACK_WEBHOOK, json={"text": msg}, timeout=5)
    except Exception as e:
        print(f"[slack] notify failed: {e}")

# ── Meraki helpers ────────────────────────────────────────────
def meraki_headers():
    return {
        "X-Cisco-Meraki-API-Key": MERAKI_API_KEY,
        "Content-Type": "application/json"
    }

def get_live_ssids():
    if not MERAKI_API_KEY or not MERAKI_NET_ID:
        return None, "Meraki not configured"
    try:
        resp = requests.get(
            f"{MERAKI_BASE}/networks/{MERAKI_NET_ID}/wireless/ssids",
            headers=meraki_headers(), timeout=10
        )
        if resp.status_code == 200:
            return resp.json(), None
        return None, f"HTTP {resp.status_code}"
    except Exception as e:
        return None, str(e)

def push_ssid_to_meraki(row, slot_index):
    """Core push logic — used by both manual push and scheduler."""
    tier_key  = str(row.get("tier") or "1")
    tier      = TIERS.get(tier_key, TIERS["1"])
    bw_kbps   = tier["kbps"]
    resp = requests.put(
        f"{MERAKI_BASE}/networks/{MERAKI_NET_ID}/wireless/ssids/{slot_index}",
        headers=meraki_headers(),
        json={
            "name": row["ssid"],
            "enabled": True,
            "authMode": "psk",
            "encryptionMode": "wpa",
            "wpaEncryptionMode": "WPA2 only",
            "psk": row["password"],
            "perClientBandwidthLimitUp":   bw_kbps,
            "perClientBandwidthLimitDown": bw_kbps,
        },
        timeout=10
    )
    return resp

# ══════════════════════════════════════════════════════════════
# SCHEDULER  (runs in background thread, checks every 60s)
# ══════════════════════════════════════════════════════════════
def scheduler_tick():
    """Called every 60s. Enables/disables SSIDs based on schedule."""
    if not MERAKI_API_KEY or not MERAKI_NET_ID:
        return
    try:
        db  = get_scheduler_db()
        cur = db.cursor()
        now = datetime.now(timezone.utc)

        # ── Auto-enable: pushed rows where enable_at has passed and not yet enabled
        cur.execute("""
            SELECT * FROM wifi_requests
            WHERE status = 'pushed'
              AND slot IS NOT NULL
              AND enable_at IS NOT NULL
              AND enable_at <= %s
              AND (schedule_status IS NULL OR schedule_status NOT IN ('enabled', 'disabled'))
        """, (now,))
        to_enable = cur.fetchall()
        for row in to_enable:
            try:
                slot_index = int(row["slot"]) - 1
                push_ssid_to_meraki(row, slot_index)
                cur.execute(
                    "UPDATE wifi_requests SET schedule_status='enabled' WHERE id=%s",
                    (row["id"],)
                )
                db.commit()
                print(f"[scheduler] Enabled SSID '{row['ssid']}' on slot {row['slot']}")
            except Exception as e:
                print(f"[scheduler] Enable error for id {row['id']}: {e}")

        # ── Auto-disable: pushed rows where disable_at has passed
        cur.execute("""
            SELECT * FROM wifi_requests
            WHERE status = 'pushed'
              AND slot IS NOT NULL
              AND disable_at IS NOT NULL
              AND disable_at <= %s
              AND (schedule_status IS NULL OR schedule_status != 'disabled')
        """, (now,))
        to_disable = cur.fetchall()
        for row in to_disable:
            try:
                slot_index = int(row["slot"]) - 1
                requests.put(
                    f"{MERAKI_BASE}/networks/{MERAKI_NET_ID}/wireless/ssids/{slot_index}",
                    headers=meraki_headers(),
                    json={"enabled": False},
                    timeout=10
                )
                cur.execute(
                    "UPDATE wifi_requests SET schedule_status='disabled' WHERE id=%s",
                    (row["id"],)
                )
                db.commit()
                print(f"[scheduler] Disabled SSID '{row['ssid']}' on slot {row['slot']}")
            except Exception as e:
                print(f"[scheduler] Disable error for id {row['id']}: {e}")

        # ── Auto-archive: pushed rows past end_date (end of day)
        cur.execute("""
            SELECT * FROM wifi_requests
            WHERE status = 'pushed'
              AND end_date IS NOT NULL
              AND (end_date::date + interval '1 day') <= NOW()
        """)
        to_archive = cur.fetchall()
        for row in to_archive:
            try:
                cur.execute(
                    "UPDATE wifi_requests SET status='archived', archived_at=NOW() WHERE id=%s",
                    (row["id"],)
                )
                db.commit()
                print(f"[scheduler] Auto-archived '{row['ssid']}' (end_date passed)")
            except Exception as e:
                print(f"[scheduler] Archive error for id {row['id']}: {e}")

        cur.close()
        db.close()
    except Exception as e:
        print(f"[scheduler] tick error: {e}")

def start_scheduler():
    def loop():
        import time
        while True:
            scheduler_tick()
            time.sleep(60)
    t = threading.Thread(target=loop, daemon=True)
    t.start()

# ══════════════════════════════════════════════════════════════
# MANAGEMENT ROUTES
# ══════════════════════════════════════════════════════════════

@app.route("/", methods=["GET"])
def index():
    return render_template("request.html", tiers=TIERS)

@app.route("/submit", methods=["POST"])
def submit():
    conf_name  = request.form.get("conf_name", "").strip()
    ssid       = request.form.get("ssid", "").strip().replace(" ", "-")
    password   = request.form.get("password", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date   = request.form.get("end_date", "").strip()
    notes      = request.form.get("notes", "").strip()
    tier       = request.form.get("tier", "1").strip()

    errors = []
    if not conf_name:      errors.append("Event name is required.")
    if not ssid:           errors.append("Network name is required.")
    if len(ssid) > 32:     errors.append("Network name must be 32 characters or less.")
    if not password:       errors.append("Password is required.")
    if not start_date or not end_date: errors.append("Start and end dates are required.")
    if start_date and end_date and end_date < start_date:
        errors.append("End date must be after start date.")
    if tier not in TIERS:  errors.append("Invalid tier selected.")

    if errors:
        return render_template("request.html", errors=errors, form=request.form, tiers=TIERS)

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "INSERT INTO wifi_requests (conf_name, ssid, password, start_date, end_date, notes, tier) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (conf_name, ssid, password, start_date, end_date, notes, tier)
    )
    db.commit()
    cur.close()

    tier_info = TIERS.get(tier, TIERS["1"])
    send_slack(
        f":wifi: *New Conference WiFi Request*\n"
        f">*Event:* {conf_name}\n"
        f">*SSID:* `{ssid}`\n"
        f">*Dates:* {start_date} → {end_date}\n"
        f">*Tier:* {tier_info['label']} ({tier_info['mbps']} Mbps)\n"
        f">*Notes:* {notes or '—'}\n"
        f">_Review it in the <https://web-production-eebee.up.railway.app/admin|IT Admin Panel>_"
    )

    return render_template("request.html", success=True, tiers=TIERS)

# ══════════════════════════════════════════════════════════════
# ADMIN ROUTES
# ══════════════════════════════════════════════════════════════

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect(url_for("admin_panel"))
        error = "Incorrect password."
    return render_template("login.html", error=error)

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))

@app.route("/admin")
@admin_required
def admin_panel():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM wifi_requests WHERE status='pending' OR status='error' ORDER BY submitted_at DESC")
    pending_list = cur.fetchall()
    cur.execute("SELECT * FROM wifi_requests WHERE status='pushed' ORDER BY pushed_at DESC")
    active_list = cur.fetchall()
    cur.close()
    live_ssids, ssid_error = get_live_ssids()
    return render_template(
        "admin.html",
        pending_list=pending_list,
        active_list=active_list,
        live_ssids=live_ssids,
        ssid_error=ssid_error,
        meraki_configured=bool(MERAKI_API_KEY and MERAKI_NET_ID),
        tiers=TIERS
    )

@app.route("/admin/ssids")
@admin_required
def api_live_ssids():
    ssids, err = get_live_ssids()
    if err:
        return jsonify({"ok": False, "error": err})
    return jsonify({"ok": True, "ssids": ssids})

@app.route("/admin/ssid/toggle", methods=["POST"])
@admin_required
def toggle_ssid():
    data    = request.json or {}
    slot    = data.get("slot")
    enabled = data.get("enabled")
    if slot is None or enabled is None:
        return jsonify({"ok": False, "error": "Missing slot or enabled"}), 400
    if not MERAKI_API_KEY or not MERAKI_NET_ID:
        return jsonify({"ok": False, "error": "Meraki not configured"}), 500
    try:
        resp = requests.put(
            f"{MERAKI_BASE}/networks/{MERAKI_NET_ID}/wireless/ssids/{slot}",
            headers=meraki_headers(),
            json={"enabled": bool(enabled)},
            timeout=10
        )
        if resp.status_code == 200:
            return jsonify({"ok": True, "slot": slot, "enabled": enabled})
        return jsonify({"ok": False, "error": f"HTTP {resp.status_code}"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/admin/push/<int:req_id>", methods=["POST"])
@admin_required
def push_ssid(req_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM wifi_requests WHERE id=%s", (req_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return jsonify({"ok": False, "error": "Request not found"}), 404

    data = request.json or {}
    slot       = data.get("slot")
    enable_at  = data.get("enable_at")
    disable_at = data.get("disable_at")
    if not slot:
        cur.close()
        return jsonify({"ok": False, "error": "No slot provided"}), 400
    if not MERAKI_API_KEY or not MERAKI_NET_ID:
        cur.close()
        return jsonify({"ok": False, "error": "Meraki not configured"}), 500

    def parse_dt(val):
        if not val: return None
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
            try: return datetime.strptime(val, fmt).replace(tzinfo=timezone.utc)
            except ValueError: continue
        return None

    ea = parse_dt(enable_at)
    da = parse_dt(disable_at)
    sched_status = "scheduled" if (ea or da) else "unscheduled"

    tier_key = str(row.get("tier") or "1")
    tier     = TIERS.get(tier_key, TIERS["1"])
    slot_index = int(slot) - 1

    try:
        resp = push_ssid_to_meraki(row, slot_index)
        if resp.status_code == 200:
            cur.execute(
                """UPDATE wifi_requests
                   SET status='pushed', slot=%s, pushed_at=NOW(), error_msg=NULL,
                       enable_at=%s, disable_at=%s, schedule_status=%s
                   WHERE id=%s""",
                (slot, ea, da, sched_status, req_id)
            )
            db.commit()
            cur.close()
            return jsonify({"ok": True, "ssid": row["ssid"], "slot": slot, "mbps": tier["mbps"]})
        else:
            err = resp.json().get("errors", [f"HTTP {resp.status_code}"])
            err_str = ", ".join(err) if isinstance(err, list) else str(err)
            cur.execute("UPDATE wifi_requests SET status='error', slot=%s, error_msg=%s WHERE id=%s", (slot, err_str, req_id))
            db.commit()
            cur.close()
            return jsonify({"ok": False, "error": err_str})
    except Exception as e:
        cur.execute("UPDATE wifi_requests SET status='error', error_msg=%s WHERE id=%s", (str(e), req_id))
        db.commit()
        cur.close()
        return jsonify({"ok": False, "error": str(e)})

@app.route("/admin/schedule/<int:req_id>", methods=["POST"])
@admin_required
def set_schedule(req_id):
    """Set or clear enable_at / disable_at for a pushed request."""
    data       = request.json or {}
    enable_at  = data.get("enable_at")   # ISO string or None
    disable_at = data.get("disable_at")  # ISO string or None

    def parse_dt(val):
        if not val:
            return None
        # Accept "2025-08-01T08:00" from datetime-local input
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    ea = parse_dt(enable_at)
    da = parse_dt(disable_at)

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE wifi_requests SET enable_at=%s, disable_at=%s, schedule_status='scheduled' WHERE id=%s",
        (ea, da, req_id)
    )
    db.commit()
    cur.close()
    return jsonify({"ok": True, "enable_at": str(ea), "disable_at": str(da)})

@app.route("/admin/delete/<int:req_id>", methods=["POST"])
@admin_required
def delete_request(req_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM wifi_requests WHERE id=%s", (req_id,))
    db.commit()
    cur.close()
    return jsonify({"ok": True})

@app.route("/admin/status")
@admin_required
def queue_status():
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM wifi_requests ORDER BY submitted_at DESC")
    rows = cur.fetchall()
    cur.close()
    return jsonify(rows)

@app.route("/admin/archive/<int:req_id>", methods=["POST"])
@admin_required
def archive_request(req_id):
    """Archive a pushed request. Optionally disable on Meraki first."""
    data           = request.json or {}
    disable_meraki = data.get("disable_meraki", False)

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM wifi_requests WHERE id=%s", (req_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return jsonify({"ok": False, "error": "Not found"}), 404

    meraki_result = None
    if disable_meraki and row["slot"] and MERAKI_API_KEY and MERAKI_NET_ID:
        try:
            resp = requests.put(
                f"{MERAKI_BASE}/networks/{MERAKI_NET_ID}/wireless/ssids/{int(row['slot']) - 1}",
                headers=meraki_headers(),
                json={"enabled": False},
                timeout=10
            )
            meraki_result = "disabled" if resp.status_code == 200 else f"error {resp.status_code}"
        except Exception as e:
            meraki_result = f"error: {str(e)}"

    cur.execute(
        "UPDATE wifi_requests SET status='archived', archived_at=NOW() WHERE id=%s",
        (req_id,)
    )
    db.commit()
    cur.close()
    return jsonify({"ok": True, "meraki": meraki_result})

@app.route("/admin/history")
@admin_required
def admin_history():
    db = get_db()
    cur = db.cursor()
    # Get distinct year-months that have pushed records
    cur.execute("""
        SELECT DISTINCT
            TO_CHAR(pushed_at, 'YYYY-MM') AS month_key,
            TO_CHAR(pushed_at, 'Month YYYY') AS month_label
        FROM wifi_requests
        WHERE status IN ('pushed','archived') AND pushed_at IS NOT NULL
        ORDER BY month_key DESC
    """)
    months = cur.fetchall()

    # Default to current month
    selected = request.args.get("month", "")
    if not selected and months:
        selected = months[0]["month_key"]

    records = []
    if selected:
        cur.execute("""
            SELECT
                id, conf_name, ssid, tier, start_date, end_date,
                notes, slot, pushed_at, disable_at, schedule_status
            FROM wifi_requests
            WHERE status IN ('pushed','archived')
              AND TO_CHAR(pushed_at, 'YYYY-MM') = %s
            ORDER BY pushed_at DESC
        """, (selected,))
        records = cur.fetchall()

    cur.close()
    return render_template(
        "history.html",
        months=months,
        selected=selected,
        records=records,
        tiers=TIERS
    )

@app.route("/admin/history/export")
@admin_required
def export_history():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    month = request.args.get("month", "")
    if not month:
        return "Month required", 400

    try:
        db = get_db()
        cur = db.cursor()
        cur.execute("""
            SELECT id, conf_name, ssid, tier, start_date, end_date,
                   notes, slot, pushed_at, status, password
            FROM wifi_requests
            WHERE status IN ('pushed','archived')
              AND TO_CHAR(pushed_at, 'YYYY-MM') = %s
            ORDER BY start_date ASC
        """, (month,))
        records = cur.fetchall()
        cur.close()
    except Exception as e:
        return f"Database error: {str(e)}", 500

    try:
        wb  = Workbook()
        ws  = wb.active
        try:
            ml = dt.strptime(month, "%Y-%m").strftime("%B %Y")
        except Exception:
            ml = month
        ws.title = f"WiFi {ml}"[:31]

        # ── Styles ──────────────────────────────────────────
        gold_fill   = PatternFill("solid", fgColor="C9A84C")
        hdr_fill    = PatternFill("solid", fgColor="1F3864")
        hdr_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        sub_fill    = PatternFill("solid", fgColor="D6E4F0")
        sub_font    = Font(name="Calibri", bold=True, color="1F3864", size=9)
        body_font   = Font(name="Calibri", size=10)
        alt_fill    = PatternFill("solid", fgColor="EBF5FB")
        white_fill  = PatternFill("solid", fgColor="FFFFFF")
        red_fill    = PatternFill("solid", fgColor="FDEDEC")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin        = Side(style="thin",   color="BDC3C7")
        thick       = Side(style="medium", color="1F3864")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_top   = Border(left=thin, right=thin, top=thick, bottom=thin)
        tier_fills  = {"1": PatternFill("solid", fgColor="D6EAF8"),
                       "2": PatternFill("solid", fgColor="D5F5E3"),
                       "3": PatternFill("solid", fgColor="FDEBD0")}
        tier_fonts  = {"1": Font(name="Calibri", size=10, bold=True, color="1A5276"),
                       "2": Font(name="Calibri", size=10, bold=True, color="1E8449"),
                       "3": Font(name="Calibri", size=10, bold=True, color="784212")}

        # ── Title ───────────────────────────────────────────
        ws.merge_cells("A1:L1")
        t = ws["A1"]
        t.value     = "Aqsarniit Hotel & Conference Centre"
        t.font      = Font(name="Calibri", bold=True, size=16, color="1F3864")
        t.alignment = left
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:L2")
        s = ws["A2"]
        s.value     = f"Wi-Fi Request Log — {ml}"
        s.font      = Font(name="Calibri", size=11, color="566573")
        s.alignment = left
        ws.row_dimensions[2].height = 18

        ws.append([])  # row 3 spacer

        # ── Sub-header (matches your spreadsheet) ───────────
        ws.merge_cells("A4:L4")
        sh = ws["A4"]
        sh.value     = "Meetings & Events Wi-Fi Requests"
        sh.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        sh.fill      = hdr_fill
        sh.alignment = center
        ws.row_dimensions[4].height = 18

        # ── Column headers ──────────────────────────────────
        headers = [
            "First Day", "Last Day", "Schedule\nStart", "Schedule\nEnd",
            "Days", "Organization / Group", "Tier",
            "SSID / Wi-Fi Name", "Password",
            "Date Requested", "Slot", "Location / Notes"
        ]
        ws.append(headers)  # row 5
        for ci in range(1, len(headers)+1):
            cell = ws.cell(row=5, column=ci)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = thick_top
        ws.row_dimensions[5].height = 30

        # ── Data rows ───────────────────────────────────────
        for i, r in enumerate(records, 1):
            row_num  = i + 5
            tier_key = str(r["tier"] or "1")
            tier_mbps = TIERS.get(tier_key, TIERS["1"])["mbps"]

            try:
                sd  = dt.strptime(str(r["start_date"]), "%Y-%m-%d")
                ed  = dt.strptime(str(r["end_date"]),   "%Y-%m-%d")
                dur = (ed - sd).days + 1
                sd_fmt = sd.strftime("%d-%b-%y")
                ed_fmt = ed.strftime("%d-%b-%y")
            except Exception:
                sd_fmt = str(r["start_date"])
                ed_fmt = str(r["end_date"])
                dur    = "—"

            try:
                req_date = r["pushed_at"].strftime("%d-%b-%y") if r["pushed_at"] else "—"
            except Exception:
                req_date = "—"

            is_archived = str(r.get("status","")) == "archived"
            base_fill   = red_fill if is_archived else (alt_fill if i % 2 == 0 else white_fill)

            row_data = [
                sd_fmt, ed_fmt,
                "7am", "6pm",           # default schedule — matches your template
                dur,
                r["conf_name"] or "",
                f"Tier {tier_key}  ({tier_mbps} Mbps)",
                r["ssid"] or "",
                r["password"] or "",
                req_date,
                r["slot"] or "",
                r["notes"] or "",
            ]
            ws.append(row_data)

            for ci, _ in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci)
                cell.font      = body_font
                cell.fill      = base_fill
                cell.border    = border
                cell.alignment = center if ci in (1,2,3,4,5,7,10,11) else left

            # Tier colour override
            tc = ws.cell(row=row_num, column=7)
            tc.fill = tier_fills.get(tier_key, white_fill)
            tc.font = tier_fonts.get(tier_key, body_font)

        # ── Summary ─────────────────────────────────────────
        sr = len(records) + 7
        ws.cell(row=sr, column=1).value = f"Total: {len(records)} request(s)"
        ws.cell(row=sr, column=1).font  = Font(name="Calibri", bold=True, size=10, color="1F3864")
        t1 = sum(1 for r in records if str(r["tier"])=="1")
        t2 = sum(1 for r in records if str(r["tier"])=="2")
        t3 = sum(1 for r in records if str(r["tier"])=="3")
        ws.merge_cells(f"B{sr}:L{sr}")
        ws.cell(row=sr, column=2).value = f"Tier 1 (25 Mbps): {t1}   |   Tier 2 (50 Mbps): {t2}   |   Tier 3 (100 Mbps): {t3}"
        ws.cell(row=sr, column=2).font  = Font(name="Calibri", size=10, color="566573")

        # ── Column widths ────────────────────────────────────
        widths = [12, 12, 9, 9, 6, 28, 18, 22, 16, 14, 6, 24]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A6"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"wifi_billing_{month}.xlsx"
        )

    except Exception as e:
        import traceback
        print(f"[export] error: {traceback.format_exc()}")
        return f"Export error: {str(e)}", 500

@app.route("/admin/history/delete/<int:req_id>", methods=["POST"])
@admin_required
def delete_history_record(req_id):
    """Hard delete a billing record — only use for mistakes/test data."""
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM wifi_requests WHERE id=%s", (req_id,))
    db.commit()
    cur.close()
    return jsonify({"ok": True})

# ── Boot ──────────────────────────────────────────────────────
init_db()
start_scheduler()

if __name__ == "__main__":
    app.run(debug=True)
